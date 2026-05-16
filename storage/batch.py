
"""
批处理计算模块 (稳定版)
从 CSV 读取多组总体参数，依次调用逻辑层各函数，输出汇总结果。
特点：
- 每步计算都捕获异常，不会因单组参数错误而中断
- 自动记录详细日志
- 结果包含输入参数、输出重量、强度校核通过状态
- 支持结果导出为 CSV 或 Excel
"""

import csv
import os
import traceback
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

# 导入已有的计算函数（调用而非复制）
from logic.loads import compute_design_loads
from logic.wing.beam import compute_beam
from logic.wing.rib import compute_ribs
from logic.wing.skin import compute_skin
from logic.other.envelope import compute_envelope
from logic.other.gear import compute_gear
from logic.other.cargo import compute_cargo_structure
from logic.other.extra import compute_extra

G = 9.81

def safe_float(value, default=0.0):
    """安全转换为浮点数，转换失败返回默认值"""
    try:
        return float(value)
    except (ValueError, TypeError):
        return default

def safe_int(value, default=0):
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default

def run_batch_from_csv(input_path, output_path, log_path='batch.log'):
    """
    从 CSV 文件读取多组总体参数，依次计算并输出结果 CSV。
    参数:
        input_path: 输入 CSV 文件路径
        output_path: 输出 CSV 文件路径
        log_path: 日志文件路径
    返回:
        成功处理的参数组数
    """
    # 准备日志
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entries = [f"Batch started at {timestamp}"]

    # 检查输入文件
    if not os.path.exists(input_path):
        msg = f"ERROR: Input file not found: {input_path}"
        print(msg)
        with open(log_path, 'w', encoding='utf-8') as lf:
            lf.write(msg)
        return 0

    results = []
    success_count = 0
    fail_count = 0

    with open(input_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for idx, row in enumerate(reader, start=1):
            row_info = f"Row {idx}"
            try:
                # 1. 读取并校验总体参数
                WTO = safe_float(row.get('WTO', row.get('起飞重量', 0)))
                if WTO <= 0:
                    raise ValueError("WTO must be positive")
                kL = safe_float(row.get('kL', row.get('气囊升力占比', 0.5)))
                nvtol = safe_float(row.get('nvtol', row.get('垂飞过载', 3.8)))
                fs = safe_float(row.get('fs', row.get('安全系数', 1.5)))
                b = safe_float(row.get('b', row.get('展长', 0)))
                if b <= 0:
                    raise ValueError("b must be positive")
                r_env = safe_float(row.get('r_env', row.get('气囊短半轴半径', 0.25)))
                D_rotor = safe_float(row.get('D_rotor', row.get('旋翼直径', 0.35)))
                delta = safe_float(row.get('delta', row.get('安全距离', 0.08)))
                cw = safe_float(row.get('cw', row.get('弦长', 0.4)))
                if cw <= 0:
                    raise ValueError("cw must be positive")
                rho = safe_float(row.get('rho', row.get('空气密度', 1.225)))
                V_cruise = safe_float(row.get('V_cruise', row.get('巡航速度', 30.0)))
                Cm0 = safe_float(row.get('Cm0', row.get('零升力矩系数', -0.05)))
                W_payload = safe_float(row.get('W_payload', row.get('载重', 50)))
                W_cargo_width = safe_float(row.get('W_cargo_width', row.get('货舱宽', 0.3)))

                # 2. 载荷计算
                loads = compute_design_loads(
                    WTO=WTO, kL=kL, nvtol=nvtol, fs=fs,
                    b=b, r_env=r_env, D_rotor=D_rotor, delta=delta,
                    cw=cw, rho=rho, V_cruise=V_cruise, Cm0=Cm0,
                    W_payload=W_payload, W_cargo_width=W_cargo_width
                )

                # 3. 梁计算 (默认单梁碳管，尺寸从 CSV 可选，若无则用默认值)
                D_tube = safe_float(row.get('D_tube', 0.04))
                t_tube = safe_float(row.get('t_tube', 0.002))
                rho_beam = safe_float(row.get('rho_beam', 1600))
                sigma_beam = safe_float(row.get('sigma_beam', 280)) * 1e6
                beam_params = {'D': D_tube, 't': t_tube, 'rho': rho_beam, 'sigma_allow': sigma_beam}
                beam_result = compute_beam(
                    loads=loads, beam_type='single',
                    section_type_front='tube', section_type_rear=None,
                    front_params=beam_params, rear_params={},
                    k_reinforce=0.0
                )

                # 4. 翼肋 (默认值)
                tmax = safe_float(row.get('tmax', 0.06))
                n_rib = safe_int(row.get('n_rib', 12))
                t_rib = safe_float(row.get('t_rib', 0.002))
                fill_rib = safe_float(row.get('fill_rib', 0.2))
                rho_rib = safe_float(row.get('rho_rib', 120))
                E_rib = safe_float(row.get('E_rib', 4)) * 1e9
                rib_result = compute_ribs(
                    tmax=tmax, cw=cw, n_rib=n_rib, t_rib=t_rib,
                    fill_rib=fill_rib, rho_rib=rho_rib, E_rib=E_rib,
                    n_rib_rein=2, t_rib_rein=0.003, fill_rib_rein=0.25,
                    rho_rib_rein=120, V_root=loads['V_root']
                )

                # 5. 蒙皮
                t_skin = safe_float(row.get('t_skin', 0.0008))
                rho_skin = safe_float(row.get('rho_skin', 1700))
                E_skin = safe_float(row.get('E_skin', 40)) * 1e9
                tau_allow = safe_float(row.get('tau_allow', 30)) * 1e6
                skin_result = compute_skin(
                    b=b, cw=cw, tmax=tmax, t_skin=t_skin,
                    rho_skin=rho_skin, E_skin=E_skin, tau_allow=tau_allow,
                    T=loads['T_aero'], d_rib=0.3
                )

                # 6. 其他重量
                env_result = compute_envelope(safe_float(row.get('m_env', 2.0)))
                gear_result = compute_gear(WTO=WTO, k_gear=safe_float(row.get('k_gear', 0.04)))
                cargo_result = compute_cargo_structure(
                    L=safe_float(row.get('L_cargo', 0.4)),
                    W=safe_float(row.get('W_cargo', 0.3)),
                    H=safe_float(row.get('H_cargo', 0.15)),
                    t_cargo=safe_float(row.get('t_cargo', 0.002)),
                    rho_cargo=safe_float(row.get('rho_cargo', 1750))
                )
                extra_result = compute_extra(safe_float(row.get('W_extra', 5.0)))

                # 7. 汇总
                W_wing = beam_result['W_beam'] + rib_result['W_ribs_total'] + skin_result['W_skin']
                W_other = env_result['W_env'] + gear_result['W_gear'] + cargo_result['W_cargo'] + extra_result['W_extra']
                W_struct = (W_wing + W_other) * 1.2  # 余量系数 20%

                bend_ok = beam_result.get('bend_ok', False)
                shear_ok = beam_result.get('shear_ok', False)
                skin_ok = skin_result.get('shear_ok', False)

                results.append({
                    'row': idx,
                    'WTO': WTO,
                    'W_struct': W_struct,
                    'W_wing': W_wing,
                    'M_root': loads['M_root'],
                    'V_root': loads['V_root'],
                    'bend_ok': bend_ok,
                    'shear_ok': shear_ok,
                    'skin_ok': skin_ok,
                    'status': 'OK'
                })
                success_count += 1
                log_entries.append(f"Row {idx}: OK, W_struct={W_struct:.1f} N")

            except Exception as e:
                fail_count += 1
                error_detail = traceback.format_exc()
                results.append({
                    'row': idx,
                    'status': 'ERROR',
                    'error_message': str(e)
                })
                log_entries.append(f"Row {idx}: ERROR - {e}\n{error_detail}")

    # 写入输出 CSV
    fieldnames = ['row', 'WTO', 'W_struct', 'W_wing', 'M_root', 'V_root',
                  'bend_ok', 'shear_ok', 'skin_ok', 'status', 'error_message']
    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(results)

    # 写入日志
    log_entries.append(f"Finished: {success_count} success, {fail_count} fail")
    with open(log_path, 'w', encoding='utf-8') as lf:
        lf.write('\n'.join(log_entries))

    return success_count

def export_batch_to_excel(results, output_path):
    """将批处理结果列表导出为 Excel (带格式)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "批处理结果"
    headers = ['序号', '起飞重量(N)', '结构总重(N)', '机翼重量(N)', '弯矩(N·m)', '弯曲通过', '剪切通过', '蒙皮通过', '状态']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for i, res in enumerate(results, 2):
        ws.cell(row=i, column=1, value=res.get('row', i-1))
        ws.cell(row=i, column=2, value=res.get('WTO', 0))
        ws.cell(row=i, column=3, value=res.get('W_struct', 0))
        ws.cell(row=i, column=4, value=res.get('W_wing', 0))
        ws.cell(row=i, column=5, value=res.get('M_root', 0))
        ws.cell(row=i, column=6, value="是" if res.get('bend_ok') else "否")
        ws.cell(row=i, column=7, value="是" if res.get('shear_ok') else "否")
        ws.cell(row=i, column=8, value="是" if res.get('skin_ok') else "否")
        ws.cell(row=i, column=9, value=res.get('status', ''))
    wb.save(output_path)
    return output_path
