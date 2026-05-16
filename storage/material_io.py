
"""
材料数据库管理 (增强版)
支持 Excel / JSON / CSV 导入导出，数据校验，自动补全，备份恢复
"""

import json
import csv
import openpyxl
import os
import copy

# 默认材料库 (内置常用航空材料)
DEFAULT_MATERIALS = [
    {"name": "T300 碳纤维", "density": 1600, "sigma_allow": 280, "E": 230},
    {"name": "T700 碳纤维", "density": 1600, "sigma_allow": 350, "E": 230},
    {"name": "7075 铝合金", "density": 2700, "sigma_allow": 220, "E": 71},
    {"name": "2024 铝合金", "density": 2700, "sigma_allow": 180, "E": 71},
    {"name": "轻木", "density": 120, "sigma_allow": 15, "E": 4},
    {"name": "玻璃纤维", "density": 1800, "sigma_allow": 100, "E": 40},
    {"name": "碳纤维预浸料", "density": 1500, "sigma_allow": 600, "E": 180},
]

class MaterialDatabase:
    """材料数据库，支持导入、搜索、校验、备份"""
    def __init__(self, filepath=None):
        self.filepath = filepath
        self.materials = copy.deepcopy(DEFAULT_MATERIALS)
        self.history = []  # 操作历史，用于撤销
        if filepath and os.path.exists(filepath):
            self.load(filepath)

    # ---------- 文件操作 ----------
    def load(self, filepath):
        """根据扩展名自动加载"""
        ext = os.path.splitext(filepath)[1].lower()
        if ext == '.xlsx':
            self._load_excel(filepath)
        elif ext == '.json':
            self._load_json(filepath)
        elif ext == '.csv':
            self._load_csv(filepath)
        else:
            raise ValueError(f"不支持的文件格式: {ext}")
        self.filepath = filepath
        self.history.append(('load', filepath))

    def save(self, filepath=None):
        """保存到指定文件，扩展名决定格式"""
        if filepath is None:
            filepath = self.filepath
        if not filepath:
            raise ValueError("未指定保存路径")
        ext = os.path.splitext(filepath)[1].lower()
        if ext == '.xlsx':
            self._save_excel(filepath)
        elif ext == '.json':
            self._save_json(filepath)
        elif ext == '.csv':
            self._save_csv(filepath)
        else:
            raise ValueError(f"不支持的格式: {ext}")
        self.history.append(('save', filepath))

    # ---------- Excel ----------
    def _load_excel(self, filepath):
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        mats = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            mats.append({
                'name': str(row[0]).strip(),
                'density': float(row[1]) if row[1] else 0,
                'sigma_allow': float(row[2]) if row[2] else 0,
                'E': float(row[3]) if row[3] else 0
            })
        self.materials = mats

    def _save_excel(self, filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "材料库"
        ws.append(['名称', '密度(kg/m3)', '许用应力(MPa)', '弹性模量(GPa)'])
        for m in self.materials:
            ws.append([m['name'], m['density'], m['sigma_allow'], m['E']])
        wb.save(filepath)

    # ---------- JSON ----------
    def _load_json(self, filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            self.materials = json.load(f)

    def _save_json(self, filepath):
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(self.materials, f, indent=2, ensure_ascii=False)

    # ---------- CSV ----------
    def _load_csv(self, filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            self.materials = [{
                'name': r['name'],
                'density': float(r['density']),
                'sigma_allow': float(r['sigma_allow']),
                'E': float(r['E'])
            } for r in reader]

    def _save_csv(self, filepath):
        with open(filepath, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=['name','density','sigma_allow','E'])
            writer.writeheader()
            writer.writerows(self.materials)

    # ---------- 查询与编辑 ----------
    def search(self, name=None, min_density=None, max_density=None,
               min_stress=None, max_stress=None):
        """模糊搜索材料"""
        results = self.materials
        if name:
            results = [m for m in results if name.lower() in m['name'].lower()]
        if min_density is not None:
            results = [m for m in results if m['density'] >= min_density]
        if max_density is not None:
            results = [m for m in results if m['density'] <= max_density]
        if min_stress is not None:
            results = [m for m in results if m['sigma_allow'] >= min_stress]
        if max_stress is not None:
            results = [m for m in results if m['sigma_allow'] <= max_stress]
        return results

    def add_material(self, name, density, sigma_allow, E):
        """添加新材料 (去重检查)"""
        for m in self.materials:
            if m['name'] == name:
                raise ValueError(f"材料 '{name}' 已存在")
        self.materials.append({
            'name': name, 'density': density,
            'sigma_allow': sigma_allow, 'E': E
        })
        self.history.append(('add', name))

    def remove_material(self, name):
        """删除指定材料"""
        self.materials = [m for m in self.materials if m['name'] != name]
        self.history.append(('remove', name))

    def update_material(self, name, **kwargs):
        """更新材料属性"""
        for m in self.materials:
            if m['name'] == name:
                m.update(kwargs)
                break
        else:
            raise ValueError(f"材料 '{name}' 不存在")

    # ---------- 校验 ----------
    def validate(self):
        """校验所有材料数据合理性，返回错误列表"""
        errors = []
        for m in self.materials:
            if m['density'] <= 0:
                errors.append(f"{m['name']}: 密度必须为正")
            if m['sigma_allow'] <= 0:
                errors.append(f"{m['name']}: 许用应力必须为正")
            if m['E'] <= 0:
                errors.append(f"{m['name']}: 弹性模量必须为正")
        return errors

    # ---------- 备份恢复 ----------
    def backup(self, backup_dir):
        """在备份目录下创建带时间戳的备份文件"""
        import time
        os.makedirs(backup_dir, exist_ok=True)
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"materials_backup_{timestamp}.json"
        path = os.path.join(backup_dir, filename)
        self._save_json(path)
        return path

    def restore(self, backup_path):
        """从备份文件恢复"""
        self._load_json(backup_path)
        self.history.append(('restore', backup_path))

    # ---------- 统计信息 ----------
    def stats(self):
        """返回材料库统计"""
        if not self.materials:
            return {}
        densities = [m['density'] for m in self.materials]
        stresses = [m['sigma_allow'] for m in self.materials]
        return {
            'count': len(self.materials),
            'avg_density': sum(densities)/len(densities),
            'min_density': min(densities),
            'max_density': max(densities),
            'avg_stress': sum(stresses)/len(stresses),
            'min_stress': min(stresses),
            'max_stress': max(stresses)
        }

# 全局单例 (供界面调用)
_db = None
def get_database():
    global _db
    if _db is None:
        _db = MaterialDatabase()
    return _db


# ========== 兼容旧接口 ==========
def load_materials_from_excel(filepath):
    """从 Excel 导入材料 (兼容接口)"""
    db = MaterialDatabase()
    db.load(filepath)
    return db.materials

def load_materials_from_json(filepath):
    """从 JSON 导入材料 (兼容接口)"""
    db = MaterialDatabase()
    db.load(filepath)
    return db.materials

def load_materials_from_csv(filepath):
    """从 CSV 导入材料 (兼容接口)"""
    db = MaterialDatabase()
    db.load(filepath)
    return db.materials
