
"""
结果导出模块 (稳定性增强版)
支持带格式的 Excel 报告、JSON 输出、CSV 流输出
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import json
import csv
import datetime

def export_detailed_report(results, filepath):
    """导出详细报告，支持 Excel / JSON / CSV (根据扩展名)"""
    ext = filepath.split('.')[-1].lower()
    if ext == 'xlsx':
        return export_to_excel(results, filepath)
    elif ext == 'json':
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
        return f"JSON 报告已保存至 {filepath}"
    elif ext == 'csv':
        with open(filepath, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['部件', '重量(N)', '重量(kg)', '占比(%)'])
            components = results.get('components', {})
            W_struct = results.get('W_struct', 1)
            for name, weight in components.items():
                writer.writerow([name, weight, round(weight/9.81,3), round(weight/W_struct*100,2)])
        return f"CSV 报告已保存至 {filepath}"
    else:
        raise ValueError(f"不支持的导出格式: {ext}")

def export_to_excel(results, filepath):
    """导出为格式化的 Excel 报告 (增强版)"""
    wb = openpyxl.Workbook()
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)

    def write_header(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin

    # ---- 摘要 ----
    ws1 = wb.active
    ws1.title = "总体摘要"
    write_header(ws1, ["参数", "数值", "单位"])
    data = [
        ("起飞重量", results.get('WTO', ''), "N"),
        ("结构总重", results.get('W_struct', ''), "N"),
        ("剩余重量", results.get('W_remain', ''), "N"),
        ("结构占比", f"{results.get('W_struct', 0)/results.get('WTO', 1)*100:.1f}" if results.get('WTO') else '', "%"),
        ("报告生成时间", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "")
    ]
    for i, (l, v, u) in enumerate(data, 2):
        ws1.cell(row=i, column=1, value=l).border = thin
        ws1.cell(row=i, column=2, value=v).border = thin
        ws1.cell(row=i, column=3, value=u).border = thin

    # ---- 重量分解 ----
    ws2 = wb.create_sheet("重量分解")
    write_header(ws2, ["部件", "重量(N)", "重量(kg)", "占比(%)"])
    comps = results.get('components', {})
    W = results.get('W_struct', 1e9)
    for i, (n, w) in enumerate(comps.items(), 2):
        ws2.cell(row=i, column=1, value=n).border = thin
        ws2.cell(row=i, column=2, value=round(w, 2)).border = thin
        ws2.cell(row=i, column=3, value=round(w/9.81, 3)).border = thin
        ws2.cell(row=i, column=4, value=round(w/W*100, 2)).border = thin

    # ---- 强度校核 ----
    ws3 = wb.create_sheet("强度校核")
    write_header(ws3, ["项目", "计算值", "许用值", "裕度%", "通过"])
    beam = results.get('beam', {})
    if beam:
        sb = beam.get('sigma_b', 0)
        sa = beam.get('sigma_allow', 1)
        mb = (sa-sb)/sa*100 if sa>0 else 0
        ws3.append(["弯曲应力(MPa)", round(sb/1e6,2), round(sa/1e6,2), round(mb,2), "是" if beam.get('bend_ok') else "否"])
        tb = beam.get('tau_beam', 0)
        ta = beam.get('tau_allow', 1)
        mt = (ta-tb)/ta*100 if ta>0 else 0
        ws3.append(["剪切应力(MPa)", round(tb/1e6,2), round(ta/1e6,2), round(mt,2), "是" if beam.get('shear_ok') else "否"])
    skin = results.get('skin', {})
    if skin:
        ts = skin.get('tau_skin', 0)
        tc = skin.get('tau_cr', 0)
        ms = (tc-ts)/tc*100 if tc>0 else 0
        ws3.append(["蒙皮剪应力(MPa)", round(ts/1e6,2), round(tc/1e6,2), round(ms,2), "是" if skin.get('shear_ok') else "否"])

    # 调整列宽
    for ws in [ws1, ws2, ws3]:
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(filepath)
    return f"Excel 报告已保存至 {filepath}"
